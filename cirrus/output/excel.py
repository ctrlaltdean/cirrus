"""
Excel workbook generator.

Reads all CSV files from the triage/ and collection/ subdirectories of a
case folder and combines them into a single analysis.xlsx workbook at the
case root.  One sheet per CSV, with frozen headers, bold styling, and
auto-sized columns.

A "Summary" sheet is prepended when triage_report.json is present, showing
case metadata, per-user verdicts, and a per-check detail table.

openpyxl is a core dependency (listed in pyproject.toml).  If it is somehow
unavailable the function returns None and the caller should ignore the result
rather than raising.
"""

from __future__ import annotations

import csv
import json
from datetime import datetime, timezone
from pathlib import Path


# Sheet name → colour (hex, no #) for the header fill.
# Triage sheets get a teal header; collection sheets get navy.
_TRIAGE_COLOUR     = "0D7377"   # teal
_COLLECTION_COLOUR = "1F4E79"   # navy
_HEADER_TEXT       = "FFFFFF"   # white text on both
_NAVY              = "0F172A"   # title background
_SUMMARY_TITLE_TXT = "FFFFFF"

# Verdict colours (background fill hex)
_VERDICT_FILL = {"high": "FEF2F2", "warn": "FFF7ED", "clean": "F0FDF4"}
_VERDICT_TEXT = {"high": "B91C1C", "warn": "C2410C", "clean": "166534"}
_STATUS_FILL  = {"high": "FEF2F2", "warn": "FFF7ED", "clean": "F0FDF4",
                 "error": "FEF2F2", "skipped": "F8FAFC"}
_STATUS_TEXT  = {"high": "B91C1C", "warn": "C2410C", "clean": "166534",
                 "error": "B91C1C", "skipped": "64748B"}


def _sheet_name(stem: str, existing: set[str]) -> str:
    """
    Produce a unique Excel sheet name (≤31 chars) from a file stem.
    Converts underscores to spaces and title-cases for readability.
    """
    name = stem.replace("_", " ").title()[:31]
    if name not in existing:
        return name
    # Deduplicate by appending a counter
    for i in range(2, 100):
        candidate = f"{name[:28]} {i}"
        if candidate not in existing:
            return candidate
    return name  # give up — openpyxl will raise if truly duplicate


def _build_summary_sheet(wb, case_dir: Path) -> None:  # type: ignore[no-untyped-def]
    """
    Prepend a Summary sheet to *wb* using triage_report.json if available,
    or a plain cover page with case metadata if not.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    ws = wb.create_sheet(title="Summary", index=0)

    def _hfill(hex_str: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_str)

    def _cell(row: int, col: int, value: str, bold: bool = False,
              fg: str = "000000", bg: str | None = None,
              align: str = "left", wrap: bool = False) -> None:
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, color=fg)
        if bg:
            c.fill = _hfill(bg)
        c.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=wrap)

    # ── Title row ──────────────────────────────────────────────────────────
    _cell(1, 1, "CIRRUS  ·  Analysis Workbook", bold=True,
          fg=_SUMMARY_TITLE_TXT, bg=_NAVY, align="left")
    ws.row_dimensions[1].height = 24
    ws.column_dimensions["A"].width = 32

    triage_path = case_dir / "triage_report.json"
    if not triage_path.exists():
        # Plain cover page — no triage data
        _cell(3, 1, "Case:", bold=True)
        _cell(3, 2, case_dir.name)
        _cell(4, 1, "Generated:", bold=True)
        _cell(4, 2, datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"))
        ws.column_dimensions["B"].width = 40
        return

    try:
        data = json.loads(triage_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return

    # ── Metadata block ─────────────────────────────────────────────────────
    overall = data.get("overall_verdict", "clean").lower()
    ov_bg   = _VERDICT_FILL.get(overall, "F8FAFC")
    ov_fg   = _VERDICT_TEXT.get(overall, "000000")

    meta_rows = [
        ("Case:",     case_dir.name,                          "Analyst:",  data.get("analyst", "")),
        ("Tenant:",   data.get("tenant", ""),                 "Generated:", (data.get("generated_at", "") or "")[:19].replace("T", " ")),
        ("Days back:", str(data.get("days", "")),             "Overall Verdict:", overall.upper()),
    ]
    for i, (k1, v1, k2, v2) in enumerate(meta_rows, start=3):
        _cell(i, 1, k1, bold=True)
        _cell(i, 2, v1)
        _cell(i, 4, k2, bold=True)
        bg = ov_bg if k2 == "Overall Verdict:" else None
        fg = ov_fg if k2 == "Overall Verdict:" else "000000"
        _cell(i, 5, v2, bold=(k2 == "Overall Verdict:"), fg=fg, bg=bg)
        ws.row_dimensions[i].height = 16

    # ── Per-user verdict table ─────────────────────────────────────────────
    HDR_ROW = 8
    _cell(HDR_ROW - 1, 1, "User Verdicts", bold=True,
          fg=_HEADER_TEXT, bg=_TRIAGE_COLOUR)
    ws.row_dimensions[HDR_ROW - 1].height = 16

    headers = ["User", "Verdict", "Flagged Checks", "Total Flags", "Top Flags"]
    for col, h in enumerate(headers, start=1):
        _cell(HDR_ROW, col, h, bold=True, fg=_HEADER_TEXT, bg=_TRIAGE_COLOUR,
              align="center")
    ws.row_dimensions[HDR_ROW].height = 16

    row = HDR_ROW + 1
    for report in data.get("reports", []):
        verdict = (report.get("verdict") or "clean").lower()
        v_bg = _VERDICT_FILL.get(verdict, "F8FAFC")
        v_fg = _VERDICT_TEXT.get(verdict, "000000")
        checks = report.get("checks", [])
        flagged = sum(1 for c in checks if c.get("status") in ("high", "warn"))
        all_flags: list[str] = []
        for c in checks:
            all_flags.extend(c.get("flags", []))
        top_flags = ", ".join(list(dict.fromkeys(all_flags))[:5])
        if len(all_flags) > 5:
            top_flags += f" +{len(all_flags) - 5}"

        _cell(row, 1, report.get("user", ""), bg=v_bg)
        _cell(row, 2, verdict.upper(), bold=True, fg=v_fg, bg=v_bg, align="center")
        _cell(row, 3, str(flagged), bg=v_bg, align="center")
        _cell(row, 4, str(len(all_flags)), bg=v_bg, align="center")
        _cell(row, 5, top_flags, bg=v_bg, wrap=True)
        ws.row_dimensions[row].height = 16
        row += 1

    row += 1  # blank separator

    # ── Check detail table ─────────────────────────────────────────────────
    _cell(row, 1, "Check Detail", bold=True,
          fg=_HEADER_TEXT, bg=_TRIAGE_COLOUR)
    ws.row_dimensions[row].height = 16
    row += 1

    det_headers = ["User", "Check", "Status", "Summary", "Flags"]
    for col, h in enumerate(det_headers, start=1):
        _cell(row, col, h, bold=True, fg=_HEADER_TEXT, bg=_TRIAGE_COLOUR,
              align="center")
    ws.row_dimensions[row].height = 16
    row += 1

    for report in data.get("reports", []):
        upn = report.get("user", "")
        for check in report.get("checks", []):
            status = (check.get("status") or "clean").lower()
            s_bg = _STATUS_FILL.get(status, "F8FAFC")
            s_fg = _STATUS_TEXT.get(status, "000000")
            flags_str = "  ·  ".join(check.get("flags", []))
            _cell(row, 1, upn, bg=s_bg)
            _cell(row, 2, check.get("label", ""), bg=s_bg)
            _cell(row, 3, status.upper(), bold=True, fg=s_fg, bg=s_bg, align="center")
            _cell(row, 4, check.get("summary", ""), bg=s_bg, wrap=True)
            _cell(row, 5, flags_str, bg=s_bg, wrap=True)
            ws.row_dimensions[row].height = 30 if flags_str else 16
            row += 1

    # ── Column widths ──────────────────────────────────────────────────────
    col_widths = [32, 14, 10, 42, 60]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


def generate_workbook(case_dir: Path) -> Path | None:
    """
    Generate ``analysis.xlsx`` in *case_dir* from all CSV files in the
    ``triage/`` and ``collection/`` subdirectories.

    Sheet order: triage sheets first, then collection sheets, each group
    sorted alphabetically.

    Returns the path to the written workbook, or ``None`` if openpyxl is
    unavailable or there are no CSV files to include.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    # Gather (sheet_label, csv_path, header_colour) in display order
    entries: list[tuple[str, Path, str]] = []
    seen_names: set[str] = set()

    for subdir, colour in (("triage", _TRIAGE_COLOUR), ("collection", _COLLECTION_COLOUR)):
        d = case_dir / subdir
        if not d.exists():
            continue
        for csv_path in sorted(d.glob("*.csv")):
            name = _sheet_name(csv_path.stem, seen_names)
            seen_names.add(name)
            entries.append((name, csv_path, colour))

    if not entries:
        return None

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # discard the default empty sheet

    # Summary sheet is always first
    _build_summary_sheet(wb, case_dir)

    for sheet_name, csv_path, hdr_colour in entries:
        try:
            with open(csv_path, encoding="utf-8", newline="") as fh:
                rows = list(csv.reader(fh))
        except Exception:
            continue

        ws = wb.create_sheet(title=sheet_name)

        if not rows:
            # Check ran but returned no records — show a placeholder so the
            # analyst knows the check executed and was clean.
            header_font = Font(bold=True, color=_HEADER_TEXT)
            header_fill = PatternFill("solid", fgColor=hdr_colour)
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
            cell = ws.cell(row=1, column=1, value="(no records)")
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            ws.column_dimensions["A"].width = 20
            ws.row_dimensions[1].height = 18
            continue
        ws.freeze_panes = "A2"

        header_font = Font(bold=True, color=_HEADER_TEXT)
        header_fill = PatternFill("solid", fgColor=hdr_colour)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

        col_widths: list[int] = [0] * len(rows[0])

        for row_idx, row in enumerate(rows, start=1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                val_len = len(str(value))
                if col_idx <= len(col_widths):
                    col_widths[col_idx - 1] = max(col_widths[col_idx - 1], val_len)
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align

        # Apply column widths (capped at 60, minimum 8)
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(min(width + 2, 60), 8)

        # Set header row height
        ws.row_dimensions[1].height = 18

    if not wb.sheetnames:
        return None

    output_path = case_dir / "analysis.xlsx"
    wb.save(output_path)
    return output_path
